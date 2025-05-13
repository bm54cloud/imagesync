#!/usr/bin/env python3
# Script to update HWSW list file
# Connect to VPN
# Script should be run from root of forked imagesync repo https://github.com/bm54cloud/imagesync
# Run script in a python environment
# python3 -m venv ~/pyvmomi-env
# source ~/pyvmomi-env/bin/activate
# Required arguments: --input (path to most recent HWSWList file), --kubeconfig, --sheet (sheet of HWSW to update) TODO: it will create a new file every time????
# Sample command: ./hwsw-auto.py --input HWSWList_05_02_2025-auto.xlsm --kubeconfig ~/.kube/config-prod --sheet Software-SIL
# Newly discovered images that do not match to a row in Software Name will print "no match found" and can be manually added

import os
import subprocess
import yaml
from openpyxl import load_workbook
from datetime import datetime, timedelta
import argparse

YAML_PATH = "images.yaml"
START_ROW = 8 # Start writing from this row 8 (first 7 rows are HEADERS)
VERSION_COL_IDX = 5  # Update column E (Version)
FULL_IMAGE_COL_IDX = 6  # Update column F (Image Name)

# Prep images.yaml
def prepare_images_yaml(path):
    yaml_template = {
        "collection": [],
        "cosign_verifiers": [],
        "destination": {
            "registry": "127.0.0.1:5000"
        },
        "exclude": [],
        "images": [],
        "include": [],
        "source": {
            "insecure": False
        }
    }

    # Check if file exists and is non-empty
    if not os.path.exists(path) or os.stat(path).st_size == 0:
        print(f"Creating or replacing {path} with default content...")
        with open(path, "w") as f:
            yaml.dump(yaml_template, f)
    else:
        print(f"{path} already exists and is not empty.")

# Run imagesync to extract images from cluster
def run_imagesync(kubeconfig_path):
    prepare_images_yaml(YAML_PATH)

    print("Running imagesync docker container...")
    cmd = [
        "docker", "run",
        "-v", f"{os.environ['HOME']}/.docker/:/home/python/.docker/",
        "-v", f"{kubeconfig_path}:/home/python/.kube/config",
        "-v", f"{os.getcwd()}/{YAML_PATH}:/app/images.yaml",
        "--rm", "docker.io/chaospuppy/imagesync:v1.6.0",
        "-f", "/app/images.yaml", "tidy"
    ]
    subprocess.run(cmd, check=True)
    print("imagesync completed.")

# Extract image versions from image name
def extract_versions():
    print(f"Reading versions from {YAML_PATH}...")
    with open(YAML_PATH, "r") as f:
        data = yaml.safe_load(f)

    image_versions = {}
    full_image_list = []

    for image_entry in data.get("images", []):
        full_name = image_entry["name"]
        if "RELEASE" in full_name:
            print(f"Skipping image with RELEASE tag: {full_name}")
            continue
        full_image_list.append(full_name)
        if ":" in full_name:
            raw_version = full_name.split(":")[-1]
            version = raw_version.split("-")[0]  # Truncate at first hyphen
            image_name = full_name.split("/")[-1].split(":")[0]
            image_versions[image_name] = version
    print(f"Extracted {len(full_image_list)} images (excluding RELEASE entries).")
    return image_versions, full_image_list

# Get Friday's date for HWSWList file name
def get_friday_filename():
    today = datetime.today()
    days_until_friday = (4 - today.weekday()) % 7  # Monday=0, Sunday=6
    this_friday = today + timedelta(days=days_until_friday)
    formatted_date = this_friday.strftime("%m_%d_%Y")
    return f"HWSWList_{formatted_date}-auto.xlsm"

def update_excel(image_versions, full_image_list, input_path):
    # Load latest HWSWlist identified as --input
    print(f"Opening Excel file: {input_path}")
    wb = load_workbook(input_path, keep_vba=True)
    ws = wb[SHEET_NAME]

    # Build mapping of {row_number: software_name} from column D to allow matching full image names to known software names
    software_rows = {}
    for row in ws.iter_rows(min_row=START_ROW, min_col=4, max_col=4):  # Column D
        cell = row[0]
        if cell.value:
            software_name = str(cell.value).strip()
            if "#manual" in software_name.lower():
                continue # Skip manually updated rows
            software_rows[cell.row] = software_name

    # Track which rows have already been written to, to avoid duplicates
    matched_rows = set()
    updated_count = 0

    # Normalize Software Names by splitting into keywords
    def tokenize(text):
        return text.lower().replace("_", "-").replace(" ", "-").split("-")

    # Loop through all Image Names extracted from images.yaml
    for full_image in full_image_list:
        # Extract version tag from Image Name and truncate at first hyphen
        if ":" in full_image:
            raw_version = full_image.split(":")[-1]
            version = raw_version.split("-")[0]
        else:
            version = ""

        image_string = full_image.lower() # Normalize image string for comparison
        best_row = None
        best_match_count = 0 # How many tokens matched
        best_tokens = []

        # Find best-matching Software Name row
        for row_num, sw_name in software_rows.items():
            if row_num in matched_rows:
                continue  # Skip already matched rows (don't write to same row twice)

            sw_tokens = tokenize(sw_name) # Break Software Name into tokens
            match_count = sum(1 for token in sw_tokens if token in image_string)

            # Find best match with highest number of keyword hits
            if match_count > best_match_count:
                best_match_count = match_count
                best_row = row_num
                best_tokens = sw_tokens

        # If a good match was found, write Version and Image Name into that row
        if best_row and best_match_count > 0:
            ws.cell(row=best_row, column=VERSION_COL_IDX, value=version)
            ws.cell(row=best_row, column=FULL_IMAGE_COL_IDX, value=full_image)
            matched_rows.add(best_row)
            updated_count += 1
            print(f"✅ Matched: '{full_image}' → row {best_row} ('{software_rows[best_row]}'), tokens: {best_tokens}, score: {best_match_count}")
        else:
            print(f"⚠️  No good match found for image: {full_image}")

    # Save updated workbook with new name based on current week's Friday date
    new_filename = get_friday_filename()
    wb.save(new_filename)
    print(f"\n✅ Update complete. {updated_count} rows written to '{SHEET_NAME}'.")
    print(f"📁 Workbook saved as: {new_filename}")

def main():
    parser = argparse.ArgumentParser(description="Update HWSW Excel sheet with image versions.")
    parser.add_argument("--input", required=True, help="Path to existing latest HWSW Excel file")
    parser.add_argument("--kubeconfig", required=True, help="Path to the kubeconfig file for the cluster")
    parser.add_argument("--sheet", required=True, help="Name of the worksheet to update (Software-SIL, Software-CP-DP, Software-vCloud)")
    args = parser.parse_args()

    global SHEET_NAME
    SHEET_NAME = args.sheet

    run_imagesync(args.kubeconfig)
    versions, image_list = extract_versions()
    update_excel(versions, image_list, args.input)

if __name__ == "__main__":
    main()

